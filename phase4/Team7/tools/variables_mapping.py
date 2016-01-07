__author__ = 'lorenzodonini'

import re
from openpyxl import Workbook

DWORD_SIZE = 4 #Default x_86 architecture
QWORD_SIZE = 8 #Default x_84 architecture
CHAR = 'CHAR' #1 byte
INT = 'INT' #4 bytes
LONG = 'LONG' #4 bytes
DOUBLE = 'DOUBLE' #8 bytes
FLOAT = 'FLOAT'
STRING = 'STRING'
STRUCT = 'STRUCT'
ARGC = 'ARGC' #Int
ARGV = 'ARGV'
PARAMETER_PASSING = 'PARAM_PASSING'
POINTER = 'POINTER'

class ParserVariable:
    UNDEFINED_LENGTH = '?'

    def __init__(self,rawName,startAddr,type):
        self.rawName = rawName
        if (startAddr[-1] == 'h'):
            self.startAddress = startAddr[:-1]
        else:
            self.startAddress = startAddr
        self.type = type
        self.logicalName = None
        self.endAddress = None
        self.length = None
        self.expected = None

    def setEndAddress(self,endAddr):
        if (endAddr == None):
            self.length = ParserVariable.UNDEFINED_LENGTH
            self.computeExpected()
            return
        self.endAddress = endAddr
        end = int(self.endAddress,16)
        start = int(self.startAddress,16)
        self.length = abs(end - start)
        self.computeExpected()

    def getStartAddressDec(self):
        return str(int(self.startAddress,16))

    def getEndAddressDec(self):
        if (self.endAddress == None):
            return None
        return str(int(self.endAddress,16))

    def computeExpected(self):
        if (self.type == 'byte' and self.length != ParserVariable.UNDEFINED_LENGTH and self.length == 1):
            self.expected = CHAR
        elif (self.rawName == 'arg_0'):
            self.expected = ARGC
        elif (self.rawName == 'arg_4'):
            self.expected = ARGV
        else:
            if (self.length == ParserVariable.UNDEFINED_LENGTH):
                self.expected = self.concatTypes([STRUCT,STRING,PARAMETER_PASSING,POINTER])
            elif (self.length == QWORD_SIZE):
                self.expected = self.concatTypes([DOUBLE,PARAMETER_PASSING])
            elif (self.length == DWORD_SIZE):
                self.expected = self.concatTypes([INT,PARAMETER_PASSING,LONG,POINTER])
            else:
                self.expected = self.concatTypes([STRUCT,STRING,PARAMETER_PASSING,POINTER])

    def concatTypes(self,types):
        result = ""
        n = len(types)
        i = 0
        while i < n:
            if (i > 0):
                result += "/"
            result += types[i]
            i += 1
        return result



def writeLineToExcel(variable,xls,row):
    xls.cell(row=row,column=1).value = variable.rawName
    xls.cell(row=row,column=2).value = variable.getStartAddressDec()
    xls.cell(row=row,column=3).value = variable.length
    xls.cell(row=row,column=4).value = variable.type
    xls.cell(row=row,column=5).value = variable.expected

def parseVariables(data):
    variables = []
    prev_var = None
    for i in range(len(data)):
        fields = data[i].split()
        var = ParserVariable(fields[1],fields[5],fields[3])
        if (prev_var != None):
            prev_var.setEndAddress(var.startAddress)
        variables.append(var)
        prev_var = var
    if prev_var != None:
        prev_var.setEndAddress(None)
    return variables


def main():
    fd = open('variables.txt','r')
    lines = fd.readlines()
    variables = parseVariables(lines)
    wb = Workbook()
    ws1 = wb.create_sheet('Mapping',0)
    ws1.cell(row=1,column=1).value = "Raw name"
    ws1.cell(row=1,column=2).value = "Address"
    ws1.cell(row=1,column=3).value = "Length"
    ws1.cell(row=1,column=4).value = "Type"
    ws1.cell(row=1,column=5).value = "Expected usage"
    ws1.cell(row=1,column=6).value = "Logical name"
    for i in range(len(variables)):
        writeLineToExcel(variables[i],ws1,i+2)
    wb.save('variables.xlsx')

main()



